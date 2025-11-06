"""
Import data from REGISTRU UNIC DE LIVRARI Excel file
Imports: Clients, Items, and Shipments with all relationships
"""

import openpyxl
from datetime import datetime
from app import create_app
from extensions import db
from models import (Client, Item, Category, ItemType, Material, MaterialSeries,
                    Shipment, ShipmentItem, Location, User)

def safe_str(value):
    """Safely convert value to string"""
    if value is None:
        return ''
    return str(value).strip()

def safe_int(value):
    """Safely convert value to integer"""
    if value is None or value == '':
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None

def safe_float(value):
    """Safely convert value to float"""
    if value is None or value == '' or value == ' ':
        return None
    try:
        # Handle special cases
        if isinstance(value, str) and value.strip() == '':
            return None
        return float(value)
    except (ValueError, TypeError):
        return None

def safe_date(value):
    """Safely convert value to date"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    try:
        if isinstance(value, str):
            # Try parsing string dates
            return datetime.strptime(value, '%d.%m.%Y')
    except:
        pass
    return None

def import_clients_and_items(wb):
    """Import clients and items from DATABASE_FG and extract from delivery registry"""
    print("\n=== IMPORTING CLIENTS AND ITEMS ===\n")

    clients_dict = {}
    items_dict = {}

    # Get delivery registry sheet
    ws_deliveries = wb['Registru unic-CLIENTI']

    # Get items database sheet
    ws_items = wb['DATABASE_FG']

    # First, create default category, type, material for items
    category = Category.query.filter_by(code='FG').first()
    if not category:
        category = Category(code='FG', name='Finished Goods', description='Customer finished goods')
        db.session.add(category)
        db.session.flush()

    item_type = ItemType.query.filter_by(code='FG').first()
    if not item_type:
        item_type = ItemType(code='FG', name='Finished Good', category_id=category.id)
        db.session.add(item_type)
        db.session.flush()

    # Import items from DATABASE_FG sheet (starting from row 5)
    print("Importing items from DATABASE_FG...")
    items_count = 0
    for row_idx, row in enumerate(ws_items.iter_rows(min_row=5, values_only=True), start=5):
        if not row[2]:  # Part number is in column 2 (index 2)
            continue

        customer_name = safe_str(row[1])  # CUSTOMER
        part_number = safe_str(row[2])  # Part number
        part_desc = safe_str(row[3])  # Part Description
        drawing_rev = safe_str(row[4])  # Drawing Revision
        product_family = safe_str(row[5])  # Product Family
        moq = safe_int(row[6])  # MOQ
        leadtime = safe_int(row[7])  # LeadTime
        frozen_zone = safe_int(row[8])  # Frozen zone
        ppap_status = safe_str(row[9])  # Status PPAP
        mfg_leadtime = safe_int(row[11])  # Manufacturing lead time
        subcontract = safe_str(row[12])  # Subcontract

        if part_number and part_number not in items_dict:
            # Create or get client
            if customer_name and customer_name not in clients_dict:
                client = Client.query.filter_by(name=customer_name).first()
                if not client:
                    # Generate unique code
                    code = customer_name[:3].upper() + str(len(clients_dict) + 1).zfill(3)
                    client = Client(
                        code=code,
                        name=customer_name,
                        is_active=True
                    )
                    db.session.add(client)
                    db.session.flush()
                    clients_dict[customer_name] = client
                    print(f"  Created client: {customer_name} ({code})")
                else:
                    clients_dict[customer_name] = client

            # Create SKU from part number
            sku = part_number

            # Check if item exists
            item = Item.query.filter_by(sku=sku).first()
            if not item:
                item = Item(
                    sku=sku,
                    name=part_desc[:200] if part_desc else sku,
                    description=f"Product Family: {product_family}, Drawing Rev: {drawing_rev}",
                    category_id=category.id,
                    type_id=item_type.id,
                    reorder_level=moq if moq else 0,
                    is_active=ppap_status == 'approved'
                )
                db.session.add(item)
                items_dict[part_number] = item
                items_count += 1
                if items_count % 50 == 0:
                    print(f"    Imported {items_count} items...")
            else:
                items_dict[part_number] = item

    db.session.flush()
    print(f"✓ Imported {items_count} items")
    print(f"✓ Imported {len(clients_dict)} clients")

    # Now extract additional clients and items from delivery registry
    print("\nExtracting clients and items from delivery registry...")
    additional_clients = 0
    additional_items = 0

    for row_idx, row in enumerate(ws_deliveries.iter_rows(min_row=10, values_only=True), start=10):
        if not row[0]:  # CUSTOMER
            continue

        customer_name = safe_str(row[0])
        part_number = safe_str(row[4])
        part_desc = safe_str(row[5])
        drawing_rev = safe_str(row[6])
        product_family = safe_str(row[14])

        # Create or get client
        if customer_name and customer_name not in clients_dict:
            client = Client.query.filter_by(name=customer_name).first()
            if not client:
                code = customer_name[:3].upper() + str(len(clients_dict) + 1).zfill(3)
                client = Client(
                    code=code,
                    name=customer_name,
                    is_active=True
                )
                db.session.add(client)
                db.session.flush()
                additional_clients += 1
            clients_dict[customer_name] = client

        # Create item if doesn't exist
        if part_number and part_number not in items_dict:
            sku = part_number
            item = Item.query.filter_by(sku=sku).first()
            if not item:
                item = Item(
                    sku=sku,
                    name=part_desc[:200] if part_desc else sku,
                    description=f"Product Family: {product_family}, Drawing Rev: {drawing_rev}",
                    category_id=category.id,
                    type_id=item_type.id,
                    is_active=True
                )
                db.session.add(item)
                additional_items += 1
            items_dict[part_number] = item

    db.session.flush()
    print(f"✓ Added {additional_clients} additional clients from deliveries")
    print(f"✓ Added {additional_items} additional items from deliveries")

    return clients_dict, items_dict

def import_shipments(wb, clients_dict, items_dict):
    """Import shipments/deliveries from Registru unic-CLIENTI"""
    print("\n=== IMPORTING SHIPMENTS/DELIVERIES ===\n")

    ws = wb['Registru unic-CLIENTI']

    # Get default location
    location = Location.query.filter_by(code='SHIP-01').first()
    if not location:
        location = Location(code='SHIP-01', name='Shipping Area', type='shipping', is_active=True)
        db.session.add(location)
        db.session.flush()

    # Get default user (admin)
    user = User.query.filter_by(username='admin').first()
    if not user:
        print("ERROR: Admin user not found. Please create admin user first.")
        return 0

    shipments_dict = {}
    shipment_count = 0

    # Row 10 onwards contains delivery data
    for row_idx, row in enumerate(ws.iter_rows(min_row=10, values_only=True), start=10):
        # Check if row has data
        if not row[0]:  # CUSTOMER
            continue

        customer_name = safe_str(row[0])
        order_type = safe_str(row[1])
        customer_po = safe_str(row[2])
        po_date = safe_date(row[3])
        part_number = safe_str(row[4])
        quantity = safe_int(row[8])
        price_unit = safe_float(row[7])
        price_order = safe_float(row[9])
        delivery_date = safe_date(row[27])  # Dată de livrare
        aviz_number = safe_str(row[28])  # AVIZ OLSTRAL
        qty_ok = safe_int(row[31])  # OK quantity
        qty_nok = safe_int(row[33])  # NOK quantity
        sales_value = safe_float(row[32])  # Sales
        invoice_number = safe_str(row[35])  # Factură/Serie
        notes = safe_str(row[34])  # Observații
        order_finished = safe_str(row[26])  # Order Finished YES/NO

        # Skip if no delivery data
        if not delivery_date or not qty_ok or qty_ok == 0:
            continue

        # Skip if order not finished
        if order_finished.lower() not in ['yes', 'y']:
            continue

        # Get client
        client = clients_dict.get(customer_name)
        if not client:
            continue

        # Get item
        item = items_dict.get(part_number)
        if not item:
            continue

        # Create unique shipment key
        shipment_key = f"{aviz_number}_{customer_name}_{delivery_date}"

        # Get or create shipment
        if shipment_key not in shipments_dict:
            # Generate shipment number
            if not aviz_number or aviz_number == '':
                last_shipment = Shipment.query.order_by(Shipment.id.desc()).first()
                if last_shipment:
                    last_num = int(last_shipment.shipment_number.split('-')[-1])
                    shipment_number = f"SHIP-{last_num + 1:06d}"
                else:
                    shipment_number = f"SHIP-{shipment_count + 1:06d}"
            else:
                # Check if this aviz number already exists
                base_shipment_number = f"AVIZ-{aviz_number}"
                existing = Shipment.query.filter_by(shipment_number=base_shipment_number).first()
                if existing:
                    # Add suffix to make it unique
                    counter = 1
                    while True:
                        shipment_number = f"{base_shipment_number}-{counter}"
                        if not Shipment.query.filter_by(shipment_number=shipment_number).first():
                            break
                        counter += 1
                else:
                    shipment_number = base_shipment_number

            notes_text = f"Customer PO: {customer_po}"
            if invoice_number:
                notes_text += f"\nInvoice: {invoice_number}"
            if notes:
                notes_text += f"\nNotes: {notes}"

            shipment = Shipment(
                shipment_number=shipment_number,
                client_id=client.id,
                from_location_id=location.id,
                ship_date=delivery_date,
                status='delivered',  # Since these are completed deliveries
                created_by=user.id,
                notes=notes_text if notes_text else None
            )
            db.session.add(shipment)
            db.session.flush()

            shipments_dict[shipment_key] = shipment
            shipment_count += 1

            if shipment_count % 100 == 0:
                print(f"    Imported {shipment_count} shipments...")
        else:
            shipment = shipments_dict[shipment_key]

        # Add shipment item
        item_notes = None
        if price_unit:
            item_notes = f"Unit price: €{price_unit:.2f}"
        if sales_value:
            if item_notes:
                item_notes += f" | Total: €{sales_value:.2f}"
            else:
                item_notes = f"Total: €{sales_value:.2f}"

        shipment_item = ShipmentItem(
            shipment_id=shipment.id,
            item_id=item.id,
            quantity=qty_ok,
            notes=item_notes
        )
        db.session.add(shipment_item)

    db.session.flush()
    print(f"✓ Imported {shipment_count} shipments")

    return shipment_count

def main():
    """Main import function"""
    print("=" * 60)
    print("IMPORTING DATA FROM REGISTRU UNIC DE LIVRARI")
    print("=" * 60)

    app = create_app()

    with app.app_context():
        # Load Excel file
        print("\nLoading Excel file...")
        wb = openpyxl.load_workbook('R-L-01 REGISTRU UNIC DE LIVRARI .xlsx', data_only=True)
        print("✓ Excel file loaded")

        # Import clients and items
        clients_dict, items_dict = import_clients_and_items(wb)

        # Import shipments
        shipments_count = import_shipments(wb, clients_dict, items_dict)

        # Commit all changes
        print("\n" + "=" * 60)
        print("COMMITTING CHANGES TO DATABASE...")
        db.session.commit()
        print("✓ All data committed successfully!")

        # Print summary
        print("\n" + "=" * 60)
        print("IMPORT SUMMARY:")
        print("=" * 60)
        print(f"Clients imported:    {len(clients_dict)}")
        print(f"Items imported:      {len(items_dict)}")
        print(f"Shipments imported:  {shipments_count}")
        print("=" * 60)
        print("\n✓ IMPORT COMPLETED SUCCESSFULLY!\n")

if __name__ == '__main__':
    main()
