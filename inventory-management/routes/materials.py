"""
Materials master data management routes
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, Material, InventoryLevel
from sqlalchemy import func, or_
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import datetime

bp = Blueprint('materials', __name__)


@bp.route('/')
@login_required
def index():
    """List all materials"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    category = request.args.get('category', '', type=str)

    query = Material.query

    # Search filter
    if search:
        query = query.filter(
            or_(
                Material.name.ilike(f'%{search}%'),
                Material.description.ilike(f'%{search}%')
            )
        )

    # Category filter
    if category:
        query = query.filter(Material.category == category)

    # Get all categories for filter dropdown
    categories = db.session.query(Material.category).filter(
        Material.category.isnot(None),
        Material.category != ''
    ).distinct().order_by(Material.category).all()
    categories = [c[0] for c in categories]

    # Pagination
    pagination = query.order_by(Material.name).paginate(
        page=page, per_page=50, error_out=False
    )
    materials = pagination.items

    # Get current stock for each material
    materials_with_stock = []
    for material in materials:
        stock_qty = db.session.query(func.sum(InventoryLevel.quantity)).filter(
            InventoryLevel.material_id == material.id
        ).scalar() or 0
        materials_with_stock.append((material, stock_qty))

    return render_template('materials/index.html',
                          materials=materials_with_stock,
                          pagination=pagination,
                          search=search,
                          category=category,
                          categories=categories)


@bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    """Create new material"""
    if request.method == 'POST':
        try:
            material = Material(
                name=request.form['name'].strip(),
                description=request.form.get('description', '').strip(),
                category=request.form.get('category', '').strip(),
                unit_of_measure=request.form['unit_of_measure'].strip(),
                reorder_level=float(request.form.get('reorder_level', 0)),
                reorder_quantity=float(request.form.get('reorder_quantity', 0)),
                active=request.form.get('active') == 'on'
            )

            db.session.add(material)
            db.session.commit()

            flash(f'Material "{material.name}" created successfully!', 'success')
            return redirect(url_for('materials.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error creating material: {str(e)}', 'danger')

    # Get existing categories for datalist
    categories = db.session.query(Material.category).filter(
        Material.category.isnot(None),
        Material.category != ''
    ).distinct().order_by(Material.category).all()
    categories = [c[0] for c in categories]

    return render_template('materials/new.html', categories=categories)


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    """Edit material"""
    material = Material.query.get_or_404(id)

    if request.method == 'POST':
        try:
            material.name = request.form['name'].strip()
            material.description = request.form.get('description', '').strip()
            material.category = request.form.get('category', '').strip()
            material.unit_of_measure = request.form['unit_of_measure'].strip()
            material.reorder_level = float(request.form.get('reorder_level', 0))
            material.reorder_quantity = float(request.form.get('reorder_quantity', 0))
            material.active = request.form.get('active') == 'on'

            db.session.commit()

            flash(f'Material "{material.name}" updated successfully!', 'success')
            return redirect(url_for('materials.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating material: {str(e)}', 'danger')

    # Get existing categories for datalist
    categories = db.session.query(Material.category).filter(
        Material.category.isnot(None),
        Material.category != ''
    ).distinct().order_by(Material.category).all()
    categories = [c[0] for c in categories]

    return render_template('materials/edit.html', material=material, categories=categories)


@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    """Delete material"""
    material = Material.query.get_or_404(id)

    # Check if material has inventory
    has_inventory = InventoryLevel.query.filter_by(material_id=id).first() is not None

    if has_inventory:
        flash(f'Cannot delete material "{material.name}" - it has inventory records.', 'danger')
        return redirect(url_for('materials.index'))

    try:
        db.session.delete(material)
        db.session.commit()
        flash(f'Material "{material.name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting material: {str(e)}', 'danger')

    return redirect(url_for('materials.index'))


@bp.route('/export')
@login_required
def export():
    """Export materials to Excel"""
    materials = Material.query.order_by(Material.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"

    # Headers
    headers = ['Name', 'Description', 'Category', 'Unit of Measure',
               'Reorder Level', 'Reorder Quantity', 'Active']
    ws.append(headers)

    # Data
    for material in materials:
        ws.append([
            material.name,
            material.description,
            material.category,
            material.unit_of_measure,
            material.reorder_level,
            material.reorder_quantity,
            'Yes' if material.active else 'No'
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"materials_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_data():
    """Import materials from Excel"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded.', 'danger')
            return redirect(url_for('materials.import_data'))

        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('materials.import_data'))

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(url_for('materials.import_data'))

        try:
            wb = load_workbook(file)
            ws = wb.active

            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            imported = 0
            updated = 0
            errors = []

            for idx, row in enumerate(rows, start=2):
                if not row[0]:  # Skip empty rows
                    continue

                try:
                    name = str(row[0]).strip()
                    description = str(row[1] or '').strip()
                    category = str(row[2] or '').strip()
                    unit_of_measure = str(row[3] or 'PCS').strip()
                    reorder_level = float(row[4] or 0)
                    reorder_quantity = float(row[5] or 0)
                    active = str(row[6] or 'Yes').lower() in ['yes', 'true', '1', 'active']

                    # Check if material exists
                    material = Material.query.filter_by(name=name).first()

                    if material:
                        # Update existing
                        material.description = description
                        material.category = category
                        material.unit_of_measure = unit_of_measure
                        material.reorder_level = reorder_level
                        material.reorder_quantity = reorder_quantity
                        material.active = active
                        updated += 1
                    else:
                        # Create new
                        material = Material(
                            name=name,
                            description=description,
                            category=category,
                            unit_of_measure=unit_of_measure,
                            reorder_level=reorder_level,
                            reorder_quantity=reorder_quantity,
                            active=active
                        )
                        db.session.add(material)
                        imported += 1

                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")

            db.session.commit()

            if errors:
                flash(f'Import completed with errors. Imported: {imported}, Updated: {updated}. Errors: {"; ".join(errors[:5])}', 'warning')
            else:
                flash(f'Import successful! Imported: {imported}, Updated: {updated}.', 'success')

            return redirect(url_for('materials.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error importing file: {str(e)}', 'danger')

    return render_template('materials/import.html')


@bp.route('/template')
@login_required
def download_template():
    """Download Excel template for materials import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials Template"

    # Headers
    headers = ['Name', 'Description', 'Category', 'Unit of Measure',
               'Reorder Level', 'Reorder Quantity', 'Active']
    ws.append(headers)

    # Sample data
    ws.append(['Steel Plate 10mm', 'Steel plate 10mm thickness', 'Metals', 'PCS', 100, 200, 'Yes'])
    ws.append(['Plastic Resin', 'High-density plastic resin', 'Plastics', 'KG', 500, 1000, 'Yes'])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='materials_template.xlsx')
