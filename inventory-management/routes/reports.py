"""
Reports routes - Inventory reports with Excel export
"""
from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from models import (db, Material, Item, Location, Bin, InventoryLevel, Batch,
                   InventoryTransaction)
from sqlalchemy import func, case
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime

bp = Blueprint('reports', __name__)


@bp.route('/')
@login_required
def index():
    """Reports dashboard"""
    return render_template('reports/index.html')


@bp.route('/stock-by-location')
@login_required
def stock_by_location():
    """Stock levels by location report"""
    location_filter = request.args.get('location', '', type=str)

    query = db.session.query(
        Location.code.label('location_code'),
        Location.name.label('location_name'),
        Bin.bin_code,
        Material.name.label('material_name'),
        Item.name.label('item_name'),
        InventoryLevel.quantity
    ).select_from(InventoryLevel).join(
        Location, InventoryLevel.location_id == Location.id
    ).outerjoin(
        Bin, InventoryLevel.bin_id == Bin.id
    ).outerjoin(
        Material, InventoryLevel.material_id == Material.id
    ).outerjoin(
        Item, InventoryLevel.item_id == Item.id
    ).filter(
        InventoryLevel.quantity > 0
    )

    if location_filter:
        query = query.filter(Location.code == location_filter)

    stock_data = query.order_by(
        Location.code, Bin.bin_code, Material.name, Item.name
    ).all()

    # Get locations for filter
    locations = Location.query.filter_by(active=True).order_by(Location.code).all()

    return render_template('reports/stock_by_location.html',
                          stock_data=stock_data,
                          locations=locations,
                          location_filter=location_filter)


@bp.route('/stock-by-location/export')
@login_required
def export_stock_by_location():
    """Export stock by location to Excel"""
    location_filter = request.args.get('location', '', type=str)

    query = db.session.query(
        Location.code.label('location_code'),
        Location.name.label('location_name'),
        Bin.bin_code,
        Material.name.label('material_name'),
        Item.name.label('item_name'),
        InventoryLevel.quantity
    ).select_from(InventoryLevel).join(
        Location, InventoryLevel.location_id == Location.id
    ).outerjoin(
        Bin, InventoryLevel.bin_id == Bin.id
    ).outerjoin(
        Material, InventoryLevel.material_id == Material.id
    ).outerjoin(
        Item, InventoryLevel.item_id == Item.id
    ).filter(
        InventoryLevel.quantity > 0
    )

    if location_filter:
        query = query.filter(Location.code == location_filter)

    stock_data = query.order_by(
        Location.code, Bin.bin_code, Material.name, Item.name
    ).all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock by Location"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = ['Location Code', 'Location Name', 'Bin', 'Material/Item', 'Quantity']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data
    for row, data in enumerate(stock_data, 2):
        ws.cell(row=row, column=1, value=data.location_code)
        ws.cell(row=row, column=2, value=data.location_name)
        ws.cell(row=row, column=3, value=data.bin_code or '')
        ws.cell(row=row, column=4, value=data.material_name or data.item_name)
        ws.cell(row=row, column=5, value=data.quantity)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"stock_by_location_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@bp.route('/inventory-valuation')
@login_required
def inventory_valuation():
    """Inventory valuation report using FIFO costs"""

    # Get all active batches with valuation
    batches_query = db.session.query(
        Location.code.label('location_code'),
        Material.name.label('material_name'),
        Item.name.label('item_name'),
        Batch.batch_number,
        Batch.quantity_available,
        Batch.cost_per_unit,
        (Batch.quantity_available * Batch.cost_per_unit).label('total_value'),
        Batch.received_date
    ).select_from(Batch).join(
        Location, Batch.location_id == Location.id
    ).outerjoin(
        Material, Batch.material_id == Material.id
    ).outerjoin(
        Item, Batch.item_id == Item.id
    ).filter(
        Batch.status == 'active',
        Batch.quantity_available > 0
    ).order_by(
        Location.code, Material.name, Item.name, Batch.received_date
    )

    batches = batches_query.all()

    # Calculate totals
    total_quantity = sum(b.quantity_available for b in batches)
    total_value = sum(b.total_value for b in batches)

    return render_template('reports/inventory_valuation.html',
                          batches=batches,
                          total_quantity=total_quantity,
                          total_value=total_value)


@bp.route('/inventory-valuation/export')
@login_required
def export_inventory_valuation():
    """Export inventory valuation to Excel"""

    batches_query = db.session.query(
        Location.code.label('location_code'),
        Material.name.label('material_name'),
        Item.name.label('item_name'),
        Batch.batch_number,
        Batch.quantity_available,
        Batch.cost_per_unit,
        (Batch.quantity_available * Batch.cost_per_unit).label('total_value'),
        Batch.received_date
    ).select_from(Batch).join(
        Location, Batch.location_id == Location.id
    ).outerjoin(
        Material, Batch.material_id == Material.id
    ).outerjoin(
        Item, Batch.item_id == Item.id
    ).filter(
        Batch.status == 'active',
        Batch.quantity_available > 0
    ).order_by(
        Location.code, Material.name, Item.name, Batch.received_date
    )

    batches = batches_query.all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Valuation"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = ['Location', 'Material/Item', 'Batch Number', 'Quantity', 'Cost/Unit', 'Total Value', 'Received Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data
    total_value = 0
    for row, batch in enumerate(batches, 2):
        ws.cell(row=row, column=1, value=batch.location_code)
        ws.cell(row=row, column=2, value=batch.material_name or batch.item_name)
        ws.cell(row=row, column=3, value=batch.batch_number)
        ws.cell(row=row, column=4, value=batch.quantity_available)
        ws.cell(row=row, column=5, value=batch.cost_per_unit)
        ws.cell(row=row, column=6, value=batch.total_value)
        ws.cell(row=row, column=7, value=batch.received_date.strftime('%Y-%m-%d') if batch.received_date else '')
        total_value += batch.total_value

    # Total row
    total_row = len(batches) + 2
    ws.cell(row=total_row, column=5, value='TOTAL:').font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_value).font = Font(bold=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventory_valuation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@bp.route('/low-stock-alerts')
@login_required
def low_stock_alerts():
    """Low stock alerts - materials and items below reorder level"""

    # Materials below reorder level
    materials_query = db.session.query(
        Material.name,
        Material.category,
        Material.reorder_level,
        Material.reorder_quantity,
        Material.unit_of_measure,
        func.sum(InventoryLevel.quantity).label('current_quantity')
    ).join(
        InventoryLevel, Material.id == InventoryLevel.material_id
    ).filter(
        Material.active == True
    ).group_by(
        Material.id
    ).having(
        func.sum(InventoryLevel.quantity) < Material.reorder_level
    ).order_by(
        Material.name
    )

    low_stock_materials = materials_query.all()

    # Items below reorder level
    items_query = db.session.query(
        Item.name,
        Item.category,
        Item.reorder_level,
        Item.reorder_quantity,
        Item.unit_of_measure,
        func.sum(InventoryLevel.quantity).label('current_quantity')
    ).join(
        InventoryLevel, Item.id == InventoryLevel.item_id
    ).filter(
        Item.active == True
    ).group_by(
        Item.id
    ).having(
        func.sum(InventoryLevel.quantity) < Item.reorder_level
    ).order_by(
        Item.name
    )

    low_stock_items = items_query.all()

    return render_template('reports/low_stock_alerts.html',
                          low_stock_materials=low_stock_materials,
                          low_stock_items=low_stock_items)


@bp.route('/low-stock-alerts/export')
@login_required
def export_low_stock_alerts():
    """Export low stock alerts to Excel"""

    # Materials below reorder level
    materials_query = db.session.query(
        Material.name,
        Material.category,
        Material.reorder_level,
        Material.reorder_quantity,
        Material.unit_of_measure,
        func.sum(InventoryLevel.quantity).label('current_quantity')
    ).join(
        InventoryLevel, Material.id == InventoryLevel.material_id
    ).filter(
        Material.active == True
    ).group_by(
        Material.id
    ).having(
        func.sum(InventoryLevel.quantity) < Material.reorder_level
    ).order_by(
        Material.name
    )

    low_stock_materials = materials_query.all()

    # Items below reorder level
    items_query = db.session.query(
        Item.name,
        Item.category,
        Item.reorder_level,
        Item.reorder_quantity,
        Item.unit_of_measure,
        func.sum(InventoryLevel.quantity).label('current_quantity')
    ).join(
        InventoryLevel, Item.id == InventoryLevel.item_id
    ).filter(
        Item.active == True
    ).group_by(
        Item.id
    ).having(
        func.sum(InventoryLevel.quantity) < Item.reorder_level
    ).order_by(
        Item.name
    )

    low_stock_items = items_query.all()

    # Create workbook
    wb = Workbook()

    # Materials sheet
    ws1 = wb.active
    ws1.title = "Low Stock Materials"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ['Name', 'Category', 'Current Qty', 'Reorder Level', 'Reorder Qty', 'UOM']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row, material in enumerate(low_stock_materials, 2):
        ws1.cell(row=row, column=1, value=material.name)
        ws1.cell(row=row, column=2, value=material.category)
        ws1.cell(row=row, column=3, value=material.current_quantity)
        ws1.cell(row=row, column=4, value=material.reorder_level)
        ws1.cell(row=row, column=5, value=material.reorder_quantity)
        ws1.cell(row=row, column=6, value=material.unit_of_measure)

    # Items sheet
    ws2 = wb.create_sheet("Low Stock Items")

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row, item in enumerate(low_stock_items, 2):
        ws2.cell(row=row, column=1, value=item.name)
        ws2.cell(row=row, column=2, value=item.category)
        ws2.cell(row=row, column=3, value=item.current_quantity)
        ws2.cell(row=row, column=4, value=item.reorder_level)
        ws2.cell(row=row, column=5, value=item.reorder_quantity)
        ws2.cell(row=row, column=6, value=item.unit_of_measure)

    # Auto-adjust column widths for both sheets
    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"low_stock_alerts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@bp.route('/transaction-history')
@login_required
def transaction_history():
    """Inventory transaction history"""
    page = request.args.get('page', 1, type=int)
    transaction_type = request.args.get('type', '', type=str)
    start_date = request.args.get('start_date', '', type=str)
    end_date = request.args.get('end_date', '', type=str)

    query = db.session.query(
        InventoryTransaction,
        Location.code.label('location_code'),
        Material.name.label('material_name'),
        Item.name.label('item_name')
    ).join(
        Location, InventoryTransaction.location_id == Location.id
    ).outerjoin(
        Material, InventoryTransaction.material_id == Material.id
    ).outerjoin(
        Item, InventoryTransaction.item_id == Item.id
    )

    if transaction_type:
        query = query.filter(InventoryTransaction.transaction_type == transaction_type)

    if start_date:
        query = query.filter(InventoryTransaction.transaction_date >= start_date)

    if end_date:
        query = query.filter(InventoryTransaction.transaction_date <= end_date)

    pagination = query.order_by(InventoryTransaction.transaction_date.desc()).paginate(
        page=page, per_page=100, error_out=False
    )

    return render_template('reports/transaction_history.html',
                          transactions=pagination.items,
                          pagination=pagination,
                          transaction_type=transaction_type,
                          start_date=start_date,
                          end_date=end_date)
