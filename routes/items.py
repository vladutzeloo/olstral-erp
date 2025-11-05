from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import Workbook
import io
from extensions import db
from models import Item, Category, ItemType, Material, MaterialSeries, InventoryLocation, Location
from filter_utils import TableFilter

items_bp = Blueprint('items', __name__)

@items_bp.route('/')
@login_required
def index():
    # Initialize filter
    table_filter = TableFilter(Item, request.args)

    # Configure filters
    table_filter.add_filter('category_id', operator='eq')
    table_filter.add_filter('type_id', operator='eq')
    table_filter.add_filter('material_id', operator='eq')
    table_filter.add_filter('is_active', operator='eq')
    table_filter.add_search(['sku', 'name', 'description', 'neo_code'])

    # Apply filters
    query = Item.query
    query = table_filter.apply(query)
    items = query.order_by(Item.sku).all()

    # Get options for dropdowns
    categories = Category.query.order_by(Category.name).all()
    types = ItemType.query.order_by(ItemType.name).all()
    materials = Material.query.order_by(Material.name).all()

    # Filter config
    filter_config = {
        'search_fields': True,
        'selects': [
            {
                'name': 'category_id',
                'label': 'Category',
                'options': [{'value': c.id, 'label': c.name} for c in categories]
            },
            {
                'name': 'type_id',
                'label': 'Type',
                'options': [{'value': t.id, 'label': t.name} for t in types]
            },
            {
                'name': 'material_id',
                'label': 'Material',
                'options': [{'value': m.id, 'label': m.name} for m in materials]
            },
            {
                'name': 'is_active',
                'label': 'Status',
                'options': [
                    {'value': '1', 'label': 'Active'},
                    {'value': '0', 'label': 'Inactive'},
                ]
            }
        ],
        'date_ranges': [],
        'summary': table_filter.get_filter_summary()
    }

    return render_template('items/index.html',
                         items=items,
                         filter_config=filter_config,
                         current_filters=table_filter.get_active_filters())

def generate_code_from_name(name, max_length=10):
    """Generate a code from a name by taking first letters and removing special chars"""
    import re
    # Remove special characters and split into words
    words = re.sub(r'[^a-zA-Z0-9\s]', '', name).upper().split()

    if not words:
        return "GEN"

    if len(words) == 1:
        # Single word: take first max_length chars
        return words[0][:max_length]
    else:
        # Multiple words: take first letter of each word
        code = ''.join(word[0] for word in words if word)
        if len(code) < 3:
            # If too short, add more letters from first word
            code = words[0][:max_length]
        return code[:max_length]

def get_or_create_category(name):
    """Get existing category by name or create a new one"""
    category = Category.query.filter(db.func.lower(Category.name) == db.func.lower(name)).first()
    if category:
        return category

    # Generate unique code
    base_code = generate_code_from_name(name)
    code = base_code
    counter = 1
    while Category.query.filter_by(code=code).first():
        code = f"{base_code[:7]}{counter:02d}"
        counter += 1

    category = Category(code=code, name=name.strip())
    db.session.add(category)
    db.session.flush()
    return category

def get_or_create_item_type(name, category):
    """Get existing item type by name and category or create a new one"""
    item_type = ItemType.query.filter(
        db.func.lower(ItemType.name) == db.func.lower(name),
        ItemType.category_id == category.id
    ).first()
    if item_type:
        return item_type

    # Generate unique code
    base_code = generate_code_from_name(name)
    code = base_code
    counter = 1
    while ItemType.query.filter_by(code=code).first():
        code = f"{base_code[:7]}{counter:02d}"
        counter += 1

    item_type = ItemType(code=code, name=name.strip(), category_id=category.id)
    db.session.add(item_type)
    db.session.flush()
    return item_type

def get_or_create_material(name):
    """Get existing material by name or create a new one"""
    if not name or not name.strip():
        return None

    material = Material.query.filter(db.func.lower(Material.name) == db.func.lower(name)).first()
    if material:
        return material

    # Generate unique code
    base_code = generate_code_from_name(name)
    code = base_code
    counter = 1
    while Material.query.filter_by(code=code).first():
        code = f"{base_code[:7]}{counter:02d}"
        counter += 1

    material = Material(code=code, name=name.strip())
    db.session.add(material)
    db.session.flush()
    return material

@items_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    if request.method == 'POST':
        try:
            # Get or create category, type, and material
            category = get_or_create_category(request.form.get('category'))
            item_type = get_or_create_item_type(request.form.get('type'), category)
            material = get_or_create_material(request.form.get('material'))

            # Generate SKU
            sku_parts = [category.code, item_type.code]
            if material:
                sku_parts.append(material.code)

            # Get next sequence number
            base_sku = '-'.join(sku_parts)
            last_item = Item.query.filter(Item.sku.like(f'{base_sku}-%')).order_by(Item.sku.desc()).first()

            if last_item:
                last_seq = int(last_item.sku.split('-')[-1])
                seq_num = last_seq + 1
            else:
                seq_num = 1

            sku = f"{base_sku}-{seq_num:04d}"

            # Create item
            item = Item(
                sku=sku,
                neo_code=request.form.get('neo_code'),
                name=request.form.get('name'),
                description=request.form.get('description'),
                category_id=category.id,
                type_id=item_type.id,
                material_id=material.id if material else None,
                unit_of_measure=request.form.get('unit_of_measure', 'PCS'),
                diameter=float(request.form.get('diameter')) if request.form.get('diameter') else None,
                length=float(request.form.get('length')) if request.form.get('length') else None,
                width=float(request.form.get('width')) if request.form.get('width') else None,
                height=float(request.form.get('height')) if request.form.get('height') else None,
                weight_kg=float(request.form.get('weight_kg')) if request.form.get('weight_kg') else None,
                reorder_level=int(request.form.get('reorder_level', 0)),
                reorder_quantity=int(request.form.get('reorder_quantity', 0)),
                cost=float(request.form.get('cost', 0)),
                price=float(request.form.get('price', 0))
            )

            db.session.add(item)
            db.session.commit()

            flash(f'Item {sku} created successfully!', 'success')
            return redirect(url_for('items.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating item: {str(e)}', 'danger')
            return redirect(url_for('items.new'))

    # GET request - load all existing categories, types, materials for autocomplete
    categories = Category.query.order_by(Category.name).all()
    types = ItemType.query.order_by(ItemType.name).all()
    materials = Material.query.order_by(Material.name).all()
    return render_template('items/new.html', categories=categories, types=types, materials=materials)

@items_bp.route('/<int:id>')
@login_required
def view(id):
    item = Item.query.get_or_404(id)
    return render_template('items/view.html', item=item)

@items_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    item = Item.query.get_or_404(id)
    
    if request.method == 'POST':
        item.neo_code = request.form.get('neo_code')
        item.name = request.form.get('name')
        item.description = request.form.get('description')
        item.unit_of_measure = request.form.get('unit_of_measure')
        item.diameter = float(request.form.get('diameter')) if request.form.get('diameter') else None
        item.length = float(request.form.get('length')) if request.form.get('length') else None
        item.width = float(request.form.get('width')) if request.form.get('width') else None
        item.height = float(request.form.get('height')) if request.form.get('height') else None
        item.weight_kg = float(request.form.get('weight_kg')) if request.form.get('weight_kg') else None
        item.reorder_level = int(request.form.get('reorder_level', 0))
        item.reorder_quantity = int(request.form.get('reorder_quantity', 0))
        item.cost = float(request.form.get('cost', 0))
        item.price = float(request.form.get('price', 0))
        
        db.session.commit()
        
        flash(f'Item {item.sku} updated successfully!', 'success')
        return redirect(url_for('items.view', id=item.id))
    
    return render_template('items/edit.html', item=item)

@items_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_items():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded!', 'danger')
            return redirect(url_for('items.import_items'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected!', 'danger')
            return redirect(url_for('items.import_items'))
        
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            flash('Only Excel files (.xlsx, .xls) are supported!', 'danger')
            return redirect(url_for('items.import_items'))
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            imported = 0
            errors = []
            
            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    category_code, type_code, material_code, neo_code, name, description, unit_of_measure, \
                    diameter, length, width, height, weight_kg, reorder_level, reorder_quantity, cost, price = row[:16]
                    
                    # Find or create category
                    category = Category.query.filter_by(code=str(category_code).upper()).first()
                    if not category:
                        errors.append(f"Row {row_num}: Category '{category_code}' not found")
                        continue
                    
                    # Find or create type
                    item_type = ItemType.query.filter_by(code=str(type_code).upper()).first()
                    if not item_type:
                        errors.append(f"Row {row_num}: Type '{type_code}' not found")
                        continue
                    
                    # Find material if specified
                    material = None
                    if material_code:
                        material = Material.query.filter_by(code=str(material_code).upper()).first()
                        if not material:
                            errors.append(f"Row {row_num}: Material '{material_code}' not found")
                            continue
                    
                    # Generate SKU
                    sku_parts = [category.code, item_type.code]
                    if material:
                        sku_parts.append(material.code)
                    
                    base_sku = '-'.join(sku_parts)
                    last_item = Item.query.filter(Item.sku.like(f'{base_sku}-%')).order_by(Item.sku.desc()).first()
                    
                    if last_item:
                        last_seq = int(last_item.sku.split('-')[-1])
                        seq_num = last_seq + 1
                    else:
                        seq_num = 1
                    
                    sku = f"{base_sku}-{seq_num:04d}"
                    
                    # Create item
                    item = Item(
                        sku=sku,
                        neo_code=neo_code,
                        name=name,
                        description=description or '',
                        category_id=category.id,
                        type_id=item_type.id,
                        material_id=material.id if material else None,
                        unit_of_measure=unit_of_measure or 'PCS',
                        diameter=float(diameter) if diameter else None,
                        length=float(length) if length else None,
                        width=float(width) if width else None,
                        height=float(height) if height else None,
                        weight_kg=float(weight_kg) if weight_kg else None,
                        reorder_level=int(reorder_level) if reorder_level else 0,
                        reorder_quantity=int(reorder_quantity) if reorder_quantity else 0,
                        cost=float(cost) if cost else 0,
                        price=float(price) if price else 0
                    )
                    
                    db.session.add(item)
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue
            
            db.session.commit()
            
            if imported > 0:
                flash(f'Successfully imported {imported} items!', 'success')
            if errors:
                flash(f'Errors: {"; ".join(errors[:5])}', 'warning')
            
            return redirect(url_for('items.index'))
            
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'danger')
            return redirect(url_for('items.import_items'))
    
    return render_template('items/import.html')

@items_bp.route('/template')
@login_required
def download_template():
    """Generate Excel template for import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Import"
    
    # Headers
    headers = ['category_code', 'type_code', 'material_code', 'neo_code', 'name', 'description', 
               'unit_of_measure', 'diameter', 'length', 'width', 'height', 'weight_kg',
               'reorder_level', 'reorder_quantity', 'cost', 'price']
    ws.append(headers)
    
    # Example row
    example = ['RAW', 'BAR', 'SS304', 'NEO-001', 'Stainless Steel Bar', '304 grade bar', 'PCS', 
               25, 1000, '', '', 2.5, 10, 100, 15.50, 25.00]
    ws.append(example)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='items_import_template.xlsx')

@items_bp.route('/api/types/<int:category_id>')
@login_required
def get_types(category_id):
    types = ItemType.query.filter_by(category_id=category_id).all()
    return jsonify([{'id': t.id, 'code': t.code, 'name': t.name} for t in types])

@items_bp.route('/categories')
@login_required
def categories():
    categories = Category.query.all()
    return render_template('items/categories.html', categories=categories)

@items_bp.route('/categories/new', methods=['GET', 'POST'])
@login_required
def new_category():
    if request.method == 'POST':
        category = Category(
            code=request.form.get('code').upper(),
            name=request.form.get('name'),
            description=request.form.get('description')
        )
        db.session.add(category)
        db.session.commit()
        
        flash(f'Category {category.name} created successfully!', 'success')
        return redirect(url_for('items.categories'))
    
    return render_template('items/new_category.html')

@items_bp.route('/types')
@login_required
def types():
    types = ItemType.query.all()
    return render_template('items/types.html', types=types)

@items_bp.route('/types/new', methods=['GET', 'POST'])
@login_required
def new_type():
    if request.method == 'POST':
        item_type = ItemType(
            code=request.form.get('code').upper(),
            name=request.form.get('name'),
            category_id=request.form.get('category_id'),
            description=request.form.get('description')
        )
        db.session.add(item_type)
        db.session.commit()
        
        flash(f'Type {item_type.name} created successfully!', 'success')
        return redirect(url_for('items.types'))
    
    categories = Category.query.all()
    return render_template('items/new_type.html', categories=categories)

@items_bp.route('/series')
@login_required
def series():
    series = MaterialSeries.query.all()
    return render_template('items/series.html', series=series)

@items_bp.route('/series/new', methods=['GET', 'POST'])
@login_required
def new_series():
    if request.method == 'POST':
        series = MaterialSeries(
            code=request.form.get('code').upper(),
            name=request.form.get('name'),
            description=request.form.get('description')
        )
        db.session.add(series)
        db.session.commit()
        
        flash(f'Material Series {series.name} created successfully!', 'success')
        return redirect(url_for('items.series'))
    
    return render_template('items/new_series.html')

@items_bp.route('/materials')
@login_required
def materials():
    materials = Material.query.all()
    return render_template('items/materials.html', materials=materials)

@items_bp.route('/materials/import', methods=['GET', 'POST'])
@login_required
def import_materials():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected!', 'danger')
            return redirect(url_for('items.import_materials'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected!', 'danger')
            return redirect(url_for('items.import_materials'))
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            imported = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                
                try:
                    code, neo_code, name, series_code, description = row[:5]
                    
                    if not code or not name:
                        errors.append(f"Row {row_num}: Missing code or name")
                        continue
                    
                    # Check if exists
                    existing = Material.query.filter_by(code=code.upper()).first()
                    if existing:
                        errors.append(f"Row {row_num}: Material {code} already exists")
                        continue
                    
                    # Get series if provided
                    series_id = None
                    if series_code:
                        series = MaterialSeries.query.filter_by(code=series_code.upper()).first()
                        if series:
                            series_id = series.id
                    
                    material = Material(
                        code=code.upper(),
                        neo_code=neo_code,
                        name=name,
                        series_id=series_id,
                        description=description
                    )
                    
                    db.session.add(material)
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue
            
            db.session.commit()
            
            if imported > 0:
                flash(f'Successfully imported {imported} materials!', 'success')
            if errors:
                flash(f'Errors: {"; ".join(errors[:5])}', 'warning')
            
            return redirect(url_for('items.materials'))
            
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'danger')
            return redirect(url_for('items.import_materials'))
    
    return render_template('items/import_materials.html')

@items_bp.route('/materials/template')
@login_required
def download_materials_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials Import"
    
    headers = ['code', 'neo_code', 'name', 'series_code', 'description']
    ws.append(headers)
    
    example = ['SS304', 'NEO-SS304', 'Stainless Steel 304', 'SS', '304 grade stainless steel']
    ws.append(example)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='materials_import_template.xlsx')

@items_bp.route('/materials/new', methods=['GET', 'POST'])
@login_required
def new_material():
    if request.method == 'POST':
        material = Material(
            code=request.form.get('code').upper(),
            neo_code=request.form.get('neo_code'),
            name=request.form.get('name'),
            series_id=request.form.get('series_id') if request.form.get('series_id') else None,
            description=request.form.get('description')
        )
        db.session.add(material)
        db.session.commit()
        
        flash(f'Material {material.name} created successfully!', 'success')
        return redirect(url_for('items.materials'))
    
    series = MaterialSeries.query.all()
    return render_template('items/new_material.html', series=series)
